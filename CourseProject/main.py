from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.edge.service import Service
from googletrans import Translator
from timeit import default_timer
from io import BytesIO
import copy
import openpyxl
import linecache
import pandas as pd
import requests



def load_everything():
    names = get_option('data-dencom="')
    for num in range(len(names)):
        names[num] = names[num].replace(' ', '_')

    strengths_file = get_option('data-conc="')
    companies = filter_companies(get_option('data-firmtarp="'))
    ATCs = get_option('data-codatc="')
    routes_file = filter_routes(get_option('data-formafarm="'))
    actives_file=get_option('data-dci="')
    for i in range(len(names)):
        original_name = names[i]
        name = names[i]
        strengths = strengths_file[i].split('+')
        company = companies[i].replace(' ', '_')
        company = company.replace('&','')
        routes = routes_file[i].split(',')
        actives = actives_file[i].split('+')
        file_path = 'test.txt'
        for strength in strengths:
            file = open(file_path, 'a+', encoding='utf-8')
            number, measurement = seperate_dosage(strength)
            file.write(
                f"concept_{name}(*<-sc_node_not_relation;;*)\n\t<-consept_medication(*<-sc_node_not_relation;;*);;")
            file.close()
            write_atc(ATCs[i], file_path)
            file = open(file_path, 'a+', encoding='utf-8')
            file.write(f"\n\t=>nrel_main_idtf:[{name} ({strength})](*<-lang_en;;*);;\n")
            file.write(f"\t=>nrel_company:{company};;\n")
            file.write(f"\t=>nrel_countries_of_sale:...(*->country_romania;;*);;\n")
            for route in routes:
                route = route.strip(' ').replace(' ', '_')
                route = route.replace('use', 'route')
                file.write(f"\t=>nrel_route_of_administration:concept_{route}(*<-sc_node_not_relation;;*);;\n")
#            file.write(f"\t=>nrel_dosage_form:concept_{form.replace(' ', '_')}(*<-sc_node_not_relation;;*);;\n")
            file.close()
            file = open(file_path, 'a+', encoding='utf-8')
            file.write(
                f"\t=>nrel_dosage:...(*\n\t\t<-sc_node_not_relation;;\n\t\t<-concept_dosage(*<-sc_node_not_relation;;*);;")
            file.write(
                f"\n\t\t=>nrel_measurement_in_{measurement}:{number}(*<-concept_number(*<-sc_node_not_relation;;*);;*);;*);;")
            for active in actives:
                active = active.split(',')[-1]
                active = active.lower().replace(' ', '_')
                file.write(f"\n\t=>nrel_active_substances:...(*->{active}(*\n\t\t<-sc_node_not_relation;;")
                file.write(f"\n\t\t<-concept_pharmacologic_substance(*<-sc_node_not_relation;;*);;*);;*);;\n\n")
            file.close()


def filter_companies(companies):
    for num in range(len(companies)):
        companies[num] = companies[num].replace('\\&quot;', '')
        companies[num] = companies[num].replace(' - ', '_')
        companies[num] = companies[num].replace(' ', '_')
        companies[num] = companies[num].replace('.', '')
        companies[num] = companies[num].replace(',', '')
        companies[num] = companies[num].replace('(', '')
        companies[num] = companies[num].replace(')', '')
        companies[num] = companies[num].replace('/', '_')

    return companies


def filter_routes(routes_file):
    for num in range(len(routes_file)):
        routes_file[num] = routes_file[num].replace('.', '')
        routes_file[num] = routes_file[num].replace('+', '')
        routes_file[num] = routes_file[num].replace(' ', '_')

    return routes_file


def filter_actives(actives_file):
    for num in range(len(actives_file)):
        actives_file[num] = actives_file[num].replace(' - ', '_')
        actives_file[num] = actives_file[num].replace(' + ', '_')
        actives_file[num] = actives_file[num].replace(' ', '_')
        actives_file[num] = actives_file[num].replace('.', '')
        actives_file[num] = actives_file[num].replace(',', '_')
        actives_file[num] = actives_file[num].replace('+', '_')
        actives_file[num] = actives_file[num].replace('(', '')
        actives_file[num] = actives_file[num].replace(')', '')
        actives_file[num] = actives_file[num].replace('/', '_')
    for num in range(len(actives_file)):
        actives_file[num] = actives_file[num].replace('___', '_')
        actives_file[num] = actives_file[num].replace('__', '_')

    return actives_file

def write_atc(atc_code, file_path):
    atc_url = 'https://ravimiregister.ee/en/Data/XML/atc.csv'
    response = requests.get(atc_url)
    atc_tree = pd.read_csv(BytesIO(response.content), delimiter=';', encoding="unicode_escape")

    with open(file_path, 'a+') as file:
        layers = [atc_code[0], atc_code[:3], atc_code[:4], atc_code[:5], atc_code]
        for layer in layers:
            description = ''
            left = 0
            right = len(atc_tree) - 1
            while left <= right:
                mid = (left + right) // 2
                print(f"right = {right}, left = {left}, mid = {mid}")
                if atc_tree['ATC kood'][mid] == layer:
                    description = atc_tree['Nimi'][mid].lower()
                    description = translate(atc_tree['Nimi'][mid], 'et', 'en').lower().replace(' ', '_')
                    break
                elif atc_tree['ATC kood'][mid] > layer:
                    right = mid - 1
                else:
                    left = mid + 1
            if description != '':
                file.write(
                    f"\n\t<-concept_{layer}_{description.replace('-','')}(*<-sc_node_not_relation;;*);;\n\t=>nrel_atc_code:[{layer}];;")

def translate(text, origin_lan, dest_lan):
    translator = Translator()
    result = translator.translate(text, src=origin_lan, dest=dest_lan).text
    return result

def remove_dozage(drug):  # возращает имя без чисел
    i = 0
    while i < len(drug):
        if drug[i].isdigit():
            return drug[:i]
        i += 1
    return drug

def seperate_dosage(dose):
    if dose == '':
         return 'not_mentioned', 'not_mentioned'
    values = dose.split('/')
    number = ''
    measurement = ''
    for value in values:
        i = 0
        while value[i].isdigit():
            i += 1
        number += value[:i] + '_'
        measurement += value[i:] + '_'
    number = number[:len(number) - 1]
    measurement = measurement[:len(measurement) - 1]
    return number, measurement


def get_button(url):
    options = Options()
    options.page_load_strategy = 'normal'
    service = Service(
        executable_path="C:\учеба\КП\msedgedriver.exe"
    )
    driver = webdriver.Edge(options=options, service=service)
    driver.maximize_window()

    try:
        # driver.implicitly_wait(15)
        driver.get(url=url)
        for i in range(0, 1538, 1):
            find_more_element = driver.find_element(By.XPATH, f'//a[text()="{i+1}"]')
            elements = driver.find_elements(By.XPATH, '//button[@class="btn btn-primary btn-xs"]')
            with open('./cures_url_list.html', 'a', encoding='utf-8') as file:
                for k in range(len(elements)):
                        file.write(f'{elements[k].get_attribute("outerHTML")}\n')
            find_more_element.click()
    except Exception as _ex:
        print(_ex)
    finally:
        driver.close()
        driver.quit()

def get_option(sub):
    with open('./cures_url_list.html', 'r', encoding='utf-8') as file:
        info = []
        for line in file:
            start = line.find(sub)
            start = start + len(sub)
            end = line.find('"', start)
            info.append(line[start:end])
    return info


def get_span():
    options = Options()
    options.page_load_strategy = 'normal'
    service = Service(
        executable_path="C:\учеба\КП\msedgedriver.exe"
    )
    driver = webdriver.Edge(
        service=service, options=options
    )
    try:
        with open('cures_url_list.html', 'r') as file:
            links = file.readlines()
            file.close()

        for link in links:
            driver.implicitly_wait(3)
            driver.get(url=f'{link}')

            elements = driver.find_elements(By.XPATH, '//span[contains(@style,"font-size:12pt")]')
            with open('source_page.html', 'a', encoding='utf-8') as file:
                for k in range(len(elements)):
                    file.write(f'\n{elements[k].get_attribute("innerHTML")}') #Здесь можно поменять формат source_page.html
                file.close()
    except Exception as _ex:
        print(_ex)
    finally:
        driver.close()
        driver.quit()

# def filter():
#     with open('./source_page.html', 'r', encoding='utf-8') as file:
#         lines = file.readlines()
#         problem = '&nbsp;'
#         for i in range(len(lines)):
#             lines[i] = lines[i].replace(problem,'', lines[i].count('&nbsp;'))
#
#     with open('./source_page.html', 'w', encoding='utf-8') as file:
#         for line in lines:
#             file.write(f'{line}')
#         file.close()

def test(url):
    options = Options()
    options.page_load_strategy = 'normal'
    service = Service(
        executable_path="C:\учеба\КП\msedgedriver.exe"
    )
    driver = webdriver.Edge(options=options, service=service)
    driver.maximize_window()

    try:
        driver.implicitly_wait(15)
        driver.get(url=url)
        wb = openpyxl.Workbook()
        sheet = wb.active
        elements = driver.find_elements(By.XPATH, '//button[@class="btn btn-primary btn-xs"]')
        for k in range(len(elements)):
            sheet.cell(row=k + 2, column=1).value = elements[k].get_attribute("data-dencom").replace("", "")
            sheet.cell(row=k + 2, column=2).value = elements[k].get_attribute("data-dci").replace("", "")
            sheet.cell(row=k + 2, column=3).value = elements[k].get_attribute("data-formafarm").replace("", "")
            sheet.cell(row=k + 2, column=4).value = elements[k].get_attribute("data-conc").replace("", "")
            sheet.cell(row=k + 2, column=5).value = elements[k].get_attribute("data-codatc").replace("", "")
            sheet.cell(row=k + 2, column=6).value = elements[k].get_attribute("data-actter").replace("", "")
            sheet.cell(row=k + 2, column=7).value = elements[k].get_attribute("data-prescript").replace("", "")
            sheet.cell(row=k + 2, column=8).value = elements[k].get_attribute("data-ambalaj").replace("", "")
            sheet.cell(row=k + 2, column=9).value = elements[k].get_attribute("data-volumamb").replace("", "")
            sheet.cell(row=k + 2, column=10).value = elements[k].get_attribute("data-valabamb").replace("", "")
            sheet.cell(row=k + 2, column=11).value = elements[k].get_attribute("data-cim").replace("", "")
            sheet.cell(row=k + 2, column=12).value = elements[k].get_attribute("data-firmtarp").replace("", "")
            sheet.cell(row=k + 2, column=13).value = elements[k].get_attribute("data-firmtard").replace("", "")
            sheet.cell(row=k + 2, column=14).value = elements[k].get_attribute("data-nrdtamb").replace("", "")


    except Exception as _ex:
        print(_ex)
    finally:
        driver.close()
        driver.quit()

def main():
    start = default_timer()
    # get_button(url = "https://nomenclator.anm.ro/medicamente?order=dci&direction=asc")
    # get_span()
    # filter()
    # set_JSON()
    # test(url = "https://nomenclator.anm.ro/medicamente?order=dci&direction=asc&page=15")
    load_everything()
    end = default_timer()
    print(f'Executing time: {end - start}')


if __name__ == '__main__':
    main()
